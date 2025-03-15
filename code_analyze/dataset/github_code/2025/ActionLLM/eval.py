import os
# os.environ["CUDA_VISIBLE_DEVICES"] = "4"
import torch
torch.set_num_threads(4)
import torch.backends.cudnn as cudnn
import numpy as np
from pathlib import Path

import util.misc as misc
from engine import predict


from actionllm.mm_adaptation import ActionLLM
from util.action_tool import read_mapping_dict
from util.opts import get_args_parser
from collections import OrderedDict
from openpyxl import load_workbook



cuda_visible_devices = os.environ.get("CUDA_VISIBLE_DEVICES")
if cuda_visible_devices is not None:
    os.environ["CUDA_VISIBLE_DEVICES"] = cuda_visible_devices
def main(args):
    device = torch.device(args.device)

    # fix the seed for reproducibility
    seed = args.seed + misc.get_rank()
    os.environ['PYTHONHASHSEED'] = str(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

    # TY
    dataset = args.dataset
    split = args.split

    data_path = os.path.join(args.data_root, dataset)
    text_feature_path = os.path.join(args.text_feature, dataset, 'feature_class_result_res50', f"sp{split}")

    mapping_file = os.path.join(data_path, 'mapping.txt')
    actions_dict = read_mapping_dict(mapping_file)

    video_file_test_path = os.path.join(data_path, 'splits', 'test.split' + split + '.bundle')
    video_file_test = open(video_file_test_path, 'r')
    video_test_list = video_file_test.read().split('\n')[:-1]

    #Small Data Fitting
    video_file_path = os.path.join(data_path, 'splits', 'train.split' + str(1) + '.bundle')
    video_file = open(video_file_path, 'r')
    video_list = video_file.read().split('\n')[:-1]

    n_class = len(actions_dict) + 1  # 19+1
    pad_idx = n_class  # 20

    print("Predict with ", args.checkpoint_file)


    model = ActionLLM(args)

    checkpoint = torch.load(args.checkpoint_file, map_location="cpu")['model']
    model_dict = model.state_dict()

    pretrained_dict = {key: value for key, value in checkpoint.items() if(key in model_dict)}  # If the key in the checkpoint is in the madel_dict, then save it to the pretrain_dict
    model.load_state_dict(pretrained_dict, strict=False)

    model.to(device)


    res_des = dict()
    res_des['checkpoint'] = args.ck_num
    obs_perc = [0.2, 0.3] 
    for obs_p in obs_perc:
        predict(model, video_test_list, args, obs_p, n_class, actions_dict, device, data_path,res_des,text_feature_path)



    # save results in Excel
    res_des = OrderedDict(res_des)
    data = [item for pair in res_des.items() for item in pair]
    
    file_path = args.pred_file
    sheet_name = 'Sheet1'
    
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    start_row = sheet.max_row + 1
    sheet.append(data)
    
    workbook.save(file_path)




if __name__ == '__main__':
    args = get_args_parser()
    args = args.parse_args()
    if args.output_dir:
        Path(args.output_dir).mkdir(parents=True, exist_ok=True)
    main(args)
