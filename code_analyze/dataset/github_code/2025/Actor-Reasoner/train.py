import time
import openpyxl
import tools
from params import Scenario_name
from llm_agent import LLM_Agent
from memory import DrivingMemory
from matplotlib.animation import FuncAnimation
import matplotlib.animation as animation
import os
from idm_controller import IDM
from time import gmtime, strftime
import matplotlib.pyplot as plt
from bayesian_game_agent import Bayesian_Agent

Sim_times = 200
# random.seed(20)
suffix = 'llama'
def open_excel(i):
    file_dir = './train/' + Scenario_name + '/excel/' + strftime("%Y-%m-%d", gmtime()) + suffix + '/'
    file_name = file_dir + str(i) + '.xlsx'

    if not os.path.exists(file_dir):
        os.makedirs(file_dir)
    workbook = openpyxl.Workbook()
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)

    # if 'Sheet' in workbook.sheetnames:
    #     del workbook['Sheet']
    return file_name, workbook

def write_data(workbook, vehicles, llm_output, if_passed, t):
    column_names = ['t', 'x', 'y', 'v', 'acc', 'theta', 'dis2des', 'type', 'action', 'HDVintent', 'HDVstyle', 'HMI', 'if_passed']
    for vehicle in vehicles:
        sheet_name = str(vehicle.id)
        if sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(column_names)
        else:
            worksheet = workbook[sheet_name]
        state = [round(vehicle.x, 2), round(vehicle.y, 2), round(vehicle.speed, 2), round(vehicle.acc, 2), round(vehicle.heading, 2), round(vehicle.dis2des, 2), vehicle.aggressiveness, llm_output[0], llm_output[2], llm_output[3], llm_output[1], if_passed]
        row_data = [t, round(vehicle.x, 2), round(vehicle.y, 2), round(vehicle.speed, 2), round(vehicle.acc, 2), round(vehicle.heading, 2), round(vehicle.dis2des, 2), vehicle.aggressiveness, llm_output[0], llm_output[2], llm_output[3], llm_output[1], if_passed]
        worksheet.append(row_data)
        worksheet.cell(row=t + 2, column=1, value=t)
        for i, item in enumerate(state):
            worksheet.cell(row=t + 2, column=i + 2, value=item)
    return workbook

class Simulator:
    def __init__(self, case_id):
        self.cav_info, self.hdv_info = tools.initialize_vehicles()
        self.case_id = case_id
        self.agent = LLM_Agent()
        self.memory = DrivingMemory()
        self.fig, self.ax = plt.subplots(figsize=(8, 8))
        self.llm_output = ['', '', '', '']
        self.instruction_info = None
        self.retrieved_instruction_info = None
        self.file_name, self.workbook = open_excel(case_id)

    def run(self):
        " ---- option 1: show animation ---- "
        # ani = FuncAnimation(self.fig, self.update, interval=10, frames=Sim_times, blit=False, repeat=False, save_count=Sim_times)
        # plt.show()

        " ---- option 2: save as gif ---- "
        ani = FuncAnimation(self.fig, self.update, interval=10, frames=Sim_times, blit=False, repeat=False, save_count=Sim_times)
        video_dir = f'./train/{Scenario_name}/video/' + strftime("%Y-%m-%d", gmtime()) + suffix + '/'
        if not os.path.exists(video_dir):
            os.makedirs(video_dir)
        ani.save(video_dir + str(self.case_id) + '.gif', dpi=50)

        " ---- option 3: save as mp4 video ---- "
        # ani = FuncAnimation(self.fig, self.update, interval=10, frames=Sim_times, blit=False, repeat=False, save_count=Sim_times)
        # Writer = animation.writers['ffmpeg']
        # writer = Writer(fps=10, codec="h264", bitrate=-1, metadata=dict(dpi=600, artist='Me'))
        # video_dir = './video/' + strftime("%Y-%m-%d", gmtime()) + '/'
        # if not os.path.exists(video_dir):
        #     os.makedirs(video_dir)
        # ani.save(video_dir + str(self.case_id) + '.mp4', writer=writer)
        # plt.close()

    def update(self, frame):
        if tools.if_passed_conflict_point(self.cav_info, self.hdv_info):
            self.llm_output[0] = 'FASTER'
        else:
            time1 = time.time()
            self.llm_output = self.agent.llm_run(self.llm_output, self.instruction_info, self.cav_info, self.hdv_info, self.memory)

        controller = IDM(self.cav_info, self.hdv_info, self.llm_output[0])
        ego_acc = controller.cal_acceleration()
        temp_cav_info = tools.kinematic_model(self.cav_info, ego_acc)

        bayesian_agent = Bayesian_Agent(self.hdv_info, self.cav_info, action_type='discrete')
        temp_hdv_info = bayesian_agent.update_state()
        # print(temp_hdv_info.x, temp_hdv_info.y, temp_hdv_info.speed, temp_hdv_info.heading, temp_hdv_info.dis2des, temp_hdv_info.entrance, temp_hdv_info.exit)
        self.hdv_info, self.cav_info = temp_hdv_info, temp_cav_info
        tools.plot_figs(self.cav_info, self.hdv_info, self.ax, self.llm_output, self.instruction_info, self.retrieved_instruction_info)
        self.instruction_info = tools.generate_simulation_hdv_instruction(self.cav_info, self.hdv_info)
        if_passed = tools.if_passed_conflict_point(self.cav_info, self.hdv_info)
        workbook = write_data(self.workbook, [self.hdv_info, self.cav_info], self.llm_output, frame, if_passed)
        workbook.save(self.file_name)
        # print(self.instruction_info)

for case in range(50):
    sim = Simulator(case)
    sim.run()

