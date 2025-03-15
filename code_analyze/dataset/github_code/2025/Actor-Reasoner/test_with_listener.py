import time
from concurrent.futures import ThreadPoolExecutor
import tools
from params import Scenario_name
from llm_agent import LLM_Agent
from memory import DrivingMemory
from matplotlib.animation import FuncAnimation
import matplotlib.animation as animation
import os
from time import gmtime, strftime
from idm_controller import IDM
import matplotlib.pyplot as plt
from bayesian_game_agent import Bayesian_Agent
import random
import numpy as np
import openpyxl
from listener import ContinuousSpeechToText

Sim_times = 200
suffix = '(with-True_instruction)'

def open_excel(i):
    file_dir = './test/' + Scenario_name + '/excel/' + strftime("%Y-%m-%d", gmtime()) + suffix + '/'
    file_name = file_dir + str(i) + '.xlsx'

    if not os.path.exists(file_dir):
        os.makedirs(file_dir)
    workbook = openpyxl.Workbook()
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
    return file_name, workbook

def write_data(workbook, vehicles, llm_output, if_passed, t):
    column_names = ['t', 'x', 'y', 'v', 'acc', 'theta', 'dis2des', 'type', 'action', 'HDVintent', 'HDVstyle', 'HMI', 'if_passed']
    for vehicle in vehicles:
        sheet_name = str(vehicle.id)
        if sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(column_names)
        else:
            worksheet = workbook[sheet_name]
        state = [round(vehicle.x, 2), round(vehicle.y, 2), round(vehicle.speed, 2), round(vehicle.acc, 2), round(vehicle.heading, 2), round(vehicle.dis2des, 2), vehicle.aggressiveness, llm_output[0], llm_output[2], llm_output[3], llm_output[1], if_passed]
        row_data = [t, round(vehicle.x, 2), round(vehicle.y, 2), round(vehicle.speed, 2), round(vehicle.acc, 2), round(vehicle.heading, 2), round(vehicle.dis2des, 2), vehicle.aggressiveness, llm_output[0], llm_output[2], llm_output[3], llm_output[1], if_passed]
        worksheet.append(row_data)
        worksheet.cell(row=t + 2, column=1, value=t)
        for i, item in enumerate(state):
            worksheet.cell(row=t + 2, column=i + 2, value=item)
    return workbook

class Simulator:
    def __init__(self, case_id, seed):
        self.seed = seed
        random.seed(self.seed)
        np.random.seed(self.seed)
        self.cav_info, self.hdv_info = tools.initialize_vehicles()
        self.case_id = case_id
        self.agent = LLM_Agent()
        self.memory = DrivingMemory()
        self.fig, self.ax = plt.subplots(figsize=(8, 8))
        self.instruction_info = None
        self.retrieved_instruction_info = None
        self.llm_output = [None, None, None, None]
        self.stop_threads = False
        self.st = time.time()
        self.executor = ThreadPoolExecutor(max_workers=3)
        self.file_name, self.workbook = open_excel(case_id)



    def run(self):
        " ---- option 1: show animation ---- "
        ani = FuncAnimation(self.fig, self.update, interval=10, frames=Sim_times, blit=False, repeat=False, save_count=Sim_times)
        plt.show()

        " ---- option 2: save as gif ---- "
        # ani = FuncAnimation(self.fig, self.update, interval=10, frames=Sim_times, blit=False, repeat=False, save_count=Sim_times)
        # video_dir = './test/' + Scenario_name + '/video/' + strftime("%Y-%m-%d", gmtime()) + suffix + '/'
        # if not os.path.exists(video_dir):
        #     os.makedirs(video_dir)
        # ani.save(video_dir + str(self.case_id) + '.gif', dpi=50)
        # print('saved')
        # plt.close()

        " ---- option 3: save as mp4 video ---- "
        # ani = FuncAnimation(self.fig, self.update, interval=10, frames=Sim_times, blit=False, repeat=False, save_count=Sim_times)
        # Writer = animation.writers['ffmpeg']
        # writer = Writer(fps=10, codec="h264", bitrate=-1, metadata=dict(dpi=600, artist='Me'))
        # video_dir = './video/' + strftime("%Y-%m-%d", gmtime()) + '/'
        # if not os.path.exists(video_dir):
        #     os.makedirs(video_dir)
        # ani.save(video_dir + str(self.case_id) + '.mp4', writer=writer)
        # plt.close()

    def update(self, frame):
        print('This is frame:', frame, round(time.time() - self.st, 2))
        time_now = time.time()
        if frame >= Sim_times - 1:
            self.stop_threads = True
            print('shutting down')

        if tools.if_passed_conflict_point(self.cav_info, self.hdv_info):
            self.llm_output[0] = 'FASTER'
            print('no conflict anymore')
            self.stop_threads = True
        else:
            self.actor()
            self.executor.submit(self.listener)
            self.executor.submit(self.reasoner)

        llm_action, hmi_advice = self.llm_output[0], self.llm_output[1]
        controller = IDM(self.cav_info, self.hdv_info, llm_action)
        ego_acc = controller.cal_acceleration()
        temp_cav_info = tools.kinematic_model(self.cav_info, ego_acc)
        bayesian_agent = Bayesian_Agent(self.hdv_info, self.cav_info, action_type='discrete')
        temp_hdv_info = bayesian_agent.update_state()
        # temp_hdv_info = driving_simulator.update_state()  # need driving simulator or real vehicle action
        self.hdv_info, self.cav_info = temp_hdv_info, temp_cav_info
        tools.plot_figs(self.cav_info, self.hdv_info, self.ax, self.llm_output, self.instruction_info, self.retrieved_instruction_info)
        if_passed = tools.if_passed_conflict_point(self.cav_info, self.hdv_info)
        # self.instruction_info = tools.generate_simulation_hdv_instruction(self.cav_info, self.hdv_info)
        workbook = write_data(self.workbook, [self.hdv_info, self.cav_info], self.llm_output, frame, if_passed)
        workbook.save(self.file_name)
        time_end = time.time()
        if time_end - time_now < 0.1:
            time.sleep(0.1 - time_end + time_now)

    def actor(self):
        time_start = time.time()
        sce_descrip = tools.scenario_experience_generator(self.cav_info, self.hdv_info, self.llm_output, self.instruction_info)
        retrieved_memory = self.memory.retrieveMemory(query_scenario=sce_descrip, top_k=1)
        self.llm_output[0] = retrieved_memory[0][0]['final_action']
        self.retrieved_instruction_info = retrieved_memory[1][0]
        retrieve_time = round(time.time() - time_start, 2)
        # print('Retrieve memory time', retrieve_time)
        self.ax.text(20, -90, f'Fast retrieve time: {round(retrieve_time, 2)}')

    def reasoner(self):
        time_start = time.time()
        if self.stop_threads:
            return
        output = self.agent.llm_run(self.llm_output, self.instruction_info, self.cav_info, self.hdv_info, self.memory, if_train_mode=False)
        self.llm_output[1:] = output[1:]
        interference_time = round(time.time() - time_start, 2)
        self.ax.text(20, -80, f'Slow interference time: {round(interference_time, 2)}')

    def listener(self):
        """实时监听语音指令"""
        self.STT = ContinuousSpeechToText()
        time_start = time.time()
        if self.stop_threads:
            return
        self.instruction_info = self.STT.listen_and_convert()
        print('Talk&Translate time', round(time.time() - time_start, 2))
        for i in range(3):
            print("Voice instruction:", self.instruction_info)



# 运行模拟器
case_num = 100
SEED_TABLE = [_ for _ in range(case_num)]
for case in range(case_num):
    seed = SEED_TABLE[case]
    sim = Simulator(case, seed)
    sim.run()
